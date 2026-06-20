import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

INPUT  = r"C:\Users\USUARIO\Desktop\CLAUDE_CODE\PRUEBA_EXCEL_BD\BASE_DATOS_LIMPIA.xlsx"
OUTPUT = r"C:\Users\USUARIO\Desktop\CLAUDE_CODE\PRUEBA_EXCEL_BD\REPORTE_PRESENTACION.xlsx"

print("Cargando datos limpios...")
df = pd.read_excel(INPUT, sheet_name="BASE DE DATOS", engine="openpyxl")

# Rellenar vacíos para análisis
df["Completo"]  = df["Completo"].fillna("NO")
df["Ejecutado"] = df["Ejecutado"].fillna("")
df["Mes"]       = df["Mes"].fillna("SIN MES")
df["Región"]    = df["Región"].fillna("SIN REGIÓN")
df["TipoEquipo"]= df["TipoEquipo"].fillna("SIN TIPO")
df["INGRESO UNITARIO 2"] = pd.to_numeric(df["INGRESO UNITARIO 2"], errors="coerce").fillna(0)
df["GASTO"]     = pd.to_numeric(df["GASTO"], errors="coerce").fillna(0)
df["Utilidad"]  = pd.to_numeric(df["Utilidad"], errors="coerce").fillna(0)

# ── Estilos reutilizables ─────────────────────────────────────────────────────
AZUL_OSCURO  = "1F4E79"
AZUL_MEDIO   = "2E75B6"
AZUL_CLARO   = "BDD7EE"
GRIS_CLARO   = "F2F2F2"
VERDE        = "375623"
VERDE_CLARO  = "E2EFDA"
ROJO         = "C00000"
BLANCO       = "FFFFFF"
AMARILLO     = "FFF2CC"

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def header_style(cell, bg=AZUL_OSCURO, fg=BLANCO, size=10, bold=True, center=True):
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(bold=bold, color=fg, size=size, name="Calibri")
    cell.alignment = Alignment(horizontal="center" if center else "left",
                               vertical="center", wrap_text=True)
    cell.border    = thin_border()

def data_style(cell, bg=BLANCO, center=False, bold=False):
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(color="000000", size=9, name="Calibri", bold=bold)
    cell.alignment = Alignment(horizontal="center" if center else "left",
                               vertical="center")
    cell.border    = thin_border()

# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 1 — RESUMEN EJECUTIVO
# ═══════════════════════════════════════════════════════════════════════════════
from openpyxl import Workbook
wb = Workbook()
ws_resumen = wb.active
ws_resumen.title = "RESUMEN EJECUTIVO"

# Título principal
ws_resumen.merge_cells("A1:H1")
c = ws_resumen["A1"]
c.value     = "PAMP 2026 — PLAN ANUAL DE MANTENIMIENTO PREVENTIVO"
c.fill      = PatternFill("solid", fgColor=AZUL_OSCURO)
c.font      = Font(bold=True, color=BLANCO, size=16, name="Calibri")
c.alignment = Alignment(horizontal="center", vertical="center")
ws_resumen.row_dimensions[1].height = 40

ws_resumen.merge_cells("A2:H2")
c = ws_resumen["A2"]
c.value     = "BASE DE DATOS CONSOLIDADA — Datos procesados y limpios"
c.fill      = PatternFill("solid", fgColor=AZUL_MEDIO)
c.font      = Font(bold=False, color=BLANCO, size=10, name="Calibri", italic=True)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_resumen.row_dimensions[2].height = 20

ws_resumen.row_dimensions[3].height = 10

# ── KPIs ──────────────────────────────────────────────────────────────────────
total_equipos   = len(df)
completos       = (df["Completo"].str.upper() == "SI").sum()
pct_completo    = round(completos / total_equipos * 100, 1) if total_equipos else 0
total_ingreso   = df["INGRESO UNITARIO 2"].sum()
total_gasto     = df["GASTO"].sum()
total_utilidad  = df["Utilidad"].sum()
regiones        = df["Región"].nunique()
tipos_equipo    = df["TipoEquipo"].nunique()

kpis = [
    ("Total Equipos",       f"{total_equipos:,}",        AZUL_MEDIO,  BLANCO),
    ("% Completados",       f"{pct_completo}%",          VERDE,       BLANCO),
    ("Regiones",            f"{regiones}",               AZUL_OSCURO, BLANCO),
    ("Tipos de Equipo",     f"{tipos_equipo}",           AZUL_OSCURO, BLANCO),
    ("Ingreso Total (S/)",  f"S/ {total_ingreso:,.2f}",  "196B5B",    BLANCO),
    ("Gasto Total (S/)",    f"S/ {total_gasto:,.2f}",    ROJO,        BLANCO),
    ("Utilidad Total (S/)", f"S/ {total_utilidad:,.2f}", "375623",    BLANCO),
]

col_start = 1
ws_resumen.row_dimensions[4].height = 20
ws_resumen.row_dimensions[5].height = 25
ws_resumen.row_dimensions[6].height = 35
ws_resumen.row_dimensions[7].height = 20

for i, (label, value, bg, fg) in enumerate(kpis):
    col = col_start + i
    lc = ws_resumen.cell(row=5, column=col, value=label)
    lc.fill      = PatternFill("solid", fgColor=bg)
    lc.font      = Font(bold=True, color=fg, size=8, name="Calibri")
    lc.alignment = Alignment(horizontal="center", vertical="center")
    lc.border    = thin_border()

    vc = ws_resumen.cell(row=6, column=col, value=value)
    vc.fill      = PatternFill("solid", fgColor=bg)
    vc.font      = Font(bold=True, color=fg, size=13, name="Calibri")
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.border    = thin_border()
    ws_resumen.column_dimensions[get_column_letter(col)].width = 20

ws_resumen.row_dimensions[8].height = 15

# ── Tabla: Equipos por Región ─────────────────────────────────────────────────
r_start = 9
ws_resumen.cell(row=r_start, column=1, value="Equipos por Región").font = \
    Font(bold=True, size=11, color=AZUL_OSCURO, name="Calibri")
ws_resumen.merge_cells(f"A{r_start}:C{r_start}")
r_start += 1

for header, col in [("Región", 1), ("Cantidad", 2), ("% del Total", 3)]:
    c = ws_resumen.cell(row=r_start, column=col, value=header)
    header_style(c, bg=AZUL_MEDIO)
r_start += 1

region_counts = df["Región"].value_counts().reset_index()
region_counts.columns = ["Región", "Cantidad"]
for i, row in region_counts.iterrows():
    pct = round(row["Cantidad"] / total_equipos * 100, 1)
    bg  = GRIS_CLARO if i % 2 == 0 else BLANCO
    for col, val in [(1, row["Región"]), (2, row["Cantidad"]), (3, f"{pct}%")]:
        c = ws_resumen.cell(row=r_start, column=col, value=val)
        data_style(c, bg=bg, center=(col > 1))
    r_start += 1

# ── Tabla: Top Tipos de Equipo ────────────────────────────────────────────────
t_start = 9
ws_resumen.cell(row=t_start, column=5, value="Top Tipos de Equipo").font = \
    Font(bold=True, size=11, color=AZUL_OSCURO, name="Calibri")
ws_resumen.merge_cells(f"E{t_start}:G{t_start}")
t_start += 1

for header, col in [("Tipo de Equipo", 5), ("Cantidad", 6), ("Ingreso (S/)", 7)]:
    c = ws_resumen.cell(row=t_start, column=col, value=header)
    header_style(c, bg=AZUL_MEDIO)
t_start += 1

tipo_stats = df.groupby("TipoEquipo").agg(
    Cantidad=("TipoEquipo", "count"),
    Ingreso=("INGRESO UNITARIO 2", "sum")
).sort_values("Cantidad", ascending=False).head(15).reset_index()

for i, row in tipo_stats.iterrows():
    bg = GRIS_CLARO if i % 2 == 0 else BLANCO
    for col, val in [(5, row["TipoEquipo"]), (6, row["Cantidad"]),
                     (7, round(row["Ingreso"], 2))]:
        c = ws_resumen.cell(row=t_start, column=col, value=val)
        data_style(c, bg=bg, center=(col == 6))
        if col == 7:
            c.number_format = '#,##0.00'
    t_start += 1

ws_resumen.freeze_panes = "A4"

# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 2 — BASE DE DATOS LIMPIA
# ═══════════════════════════════════════════════════════════════════════════════
ws_data = wb.create_sheet("BASE DE DATOS")

# Encabezados
for col_idx, col_name in enumerate(df.columns, 1):
    c = ws_data.cell(row=1, column=col_idx, value=col_name)
    header_style(c)

ws_data.row_dimensions[1].height = 38

# Datos
for row_idx, row in df.iterrows():
    excel_row = row_idx + 2
    bg = AZUL_CLARO if row_idx % 2 == 0 else BLANCO
    for col_idx, val in enumerate(row, 1):
        v = "" if (isinstance(val, float) and np.isnan(val)) else val
        c = ws_data.cell(row=excel_row, column=col_idx, value=v)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.font      = Font(size=9, name="Calibri")
        c.alignment = Alignment(vertical="center")
        c.border    = thin_border()
        # Formato numérico para financieras
        if col_name in ["INGRESO UNITARIO 2", "INGRESO OC", "GASTO", "Utilidad"]:
            c.number_format = '#,##0.00'

# Ancho de columnas
col_widths = {
    "Orden": 7, "Orden PAM 2026": 10, "CodigoEquipo": 16,
    "ST": 14, "Inmueble": 25, "CodigoInventario": 16,
    "Nombre": 20, "TipoEquipo": 18, "NombreGrupoMantenimiento": 22,
    "NombreUnidadMantenimiento": 22, "Cod Inmueble": 12,
    "Estado del Equipo (Operativo / Inoperativo)": 28,
    "Sistema": 14, "Región": 12, "Ubicación": 12,
    "Detalle Mes": 12, "Mes": 10,
    "INGRESO UNITARIO 2": 18, "INGRESO OC": 14, "GASTO": 12,
    "Utilidad": 14, "Proveedor": 20, "Cotizado PAM": 14,
    "Delegado": 16, "Certificado": 14, "Programado": 12,
    "Ejecutado": 12, "Acta": 10, "Informe": 10,
    "C.O.": 10, "Completo": 12, "EDI": 10,
    "Fecha Inicio": 14, "Fecha Fin": 14,
    "Detalle VH": 20, "OC TG": 12,
}
for col_idx, col_name in enumerate(df.columns, 1):
    w = col_widths.get(col_name, 14)
    ws_data.column_dimensions[get_column_letter(col_idx)].width = w

ws_data.freeze_panes = "A2"
ws_data.auto_filter.ref = ws_data.dimensions

# ═══════════════════════════════════════════════════════════════════════════════
#  SHEET 3 — ANÁLISIS POR MES
# ═══════════════════════════════════════════════════════════════════════════════
ws_mes = wb.create_sheet("ANÁLISIS POR MES")

orden_meses = ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
               "JULIO","AGOSTO","SEPTIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"]
mes_stats = df.groupby("Mes").agg(
    Equipos=("Mes","count"),
    Ingreso=("INGRESO UNITARIO 2","sum"),
    Gasto=("GASTO","sum"),
    Utilidad=("Utilidad","sum")
).reindex(orden_meses, fill_value=0).reset_index()

ws_mes.merge_cells("A1:E1")
c = ws_mes["A1"]
c.value     = "ANÁLISIS FINANCIERO POR MES — PAMP 2026"
c.fill      = PatternFill("solid", fgColor=AZUL_OSCURO)
c.font      = Font(bold=True, color=BLANCO, size=14, name="Calibri")
c.alignment = Alignment(horizontal="center", vertical="center")
ws_mes.row_dimensions[1].height = 35

headers = ["Mes", "Equipos", "Ingreso (S/)", "Gasto (S/)", "Utilidad (S/)"]
for col_idx, h in enumerate(headers, 1):
    c = ws_mes.cell(row=2, column=col_idx, value=h)
    header_style(c, bg=AZUL_MEDIO)
    ws_mes.column_dimensions[get_column_letter(col_idx)].width = 18

for i, row in mes_stats.iterrows():
    r = i + 3
    bg = GRIS_CLARO if i % 2 == 0 else BLANCO
    vals = [row["Mes"], row["Equipos"],
            round(row["Ingreso"],2), round(row["Gasto"],2), round(row["Utilidad"],2)]
    for col_idx, val in enumerate(vals, 1):
        c = ws_mes.cell(row=r, column=col_idx, value=val)
        data_style(c, bg=bg, center=(col_idx <= 2))
        if col_idx > 2:
            c.number_format = '#,##0.00'
        util_val = vals[4] if col_idx == 5 else None
        if util_val is not None and util_val < 0:
            c.font = Font(color=ROJO, bold=True, size=9, name="Calibri")

ws_mes.freeze_panes = "A3"

wb.save(OUTPUT)
print(f"Reporte guardado: {OUTPUT}")
print("Hojas creadas:")
print("  1. RESUMEN EJECUTIVO — KPIs + tablas de resumen")
print("  2. BASE DE DATOS     — Datos limpios con formato")
print("  3. ANÁLISIS POR MES  — Resumen financiero mensual")
