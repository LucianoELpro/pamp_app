import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE  = r"C:\Users\USUARIO\Desktop\CLAUDE_CODE\PRUEBA_EXCEL_BD\PAMP 2026 CONSOLIDADO 6-5.xlsx"
OUTPUT_FILE = r"C:\Users\USUARIO\Desktop\CLAUDE_CODE\PRUEBA_EXCEL_BD\BASE_DATOS_LIMPIA.xlsx"
SHEET_NAME  = "BASE DE DATOS"

print("Leyendo archivo...")
df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME, engine="openpyxl")
print(f"  Shape original: {df.shape}")

# ── 1. Eliminar columnas fantasma (sin encabezado o Unnamed) ──────────────────
df = df.loc[:, ~df.columns.str.startswith("Unnamed")]
print(f"  Tras quitar columnas fantasma: {df.shape}")

# ── 2. Eliminar filas completamente vacías ────────────────────────────────────
df.dropna(how="all", inplace=True)
df.reset_index(drop=True, inplace=True)
print(f"  Tras quitar filas vacías: {df.shape}")

# ── 3. Limpiar espacios en encabezados ───────────────────────────────────────
df.columns = [c.strip() for c in df.columns]

# ── 4. Limpiar texto en columnas de tipo objeto ───────────────────────────────
text_cols = df.select_dtypes(include="object").columns
for col in text_cols:
    df[col] = (
        df[col]
        .astype(str)
        .str.strip()
        .replace({"nan": "", "None": "", "-": "", "N/A": ""})
        .replace("", np.nan)          # vacíos → NaN real para consistencia
    )

# ── 5. Normalizar columnas clave a MAYÚSCULAS ─────────────────────────────────
upper_cols = [
    "TipoEquipo", "NombreGrupoMantenimiento", "NombreUnidadMantenimiento",
    "Estado del Equipo (Operativo / Inoperativo)", "Sistema",
    "Región", "Ubicación", "Mes", "Proveedor", "Completo"
]
for col in upper_cols:
    if col in df.columns:
        df[col] = df[col].str.upper().str.strip()

# ── 6. Columnas financieras: nulos → 0 ───────────────────────────────────────
fin_cols = ["INGRESO UNITARIO 2", "INGRESO OC", "GASTO", "Utilidad",
            "Cotizado PAM"]
for col in fin_cols:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

# ── 7. Recalcular Utilidad donde sea 0 y haya ingreso/gasto ──────────────────
mask = (df["Utilidad"] == 0) & (df["INGRESO UNITARIO 2"] != 0)
df.loc[mask, "Utilidad"] = df.loc[mask, "INGRESO UNITARIO 2"] - df.loc[mask, "GASTO"]

# ── 8. Columnas binarias numéricas (1/0/NULL) → SI / NO ──────────────────────
# Programado y Ejecutado vienen como número (1 = sí, 0/NULL = no)
binary_num_cols = ["Programado", "Ejecutado", "Acta", "Informe", "C.O."]
for col in binary_num_cols:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors="coerce")
        df[col] = df[col].apply(lambda x: "SI" if x == 1 else "NO")

# Columnas de estado que ya vienen como texto
text_status_cols = ["Completo", "EDI", "Delegado", "Certificado"]
for col in text_status_cols:
    if col in df.columns and df[col].dtype == object:
        df[col] = df[col].str.upper().str.strip()

# ── 9. Fechas ─────────────────────────────────────────────────────────────────
for date_col in ["Fecha Inicio", "Fecha Fin"]:
    if date_col in df.columns:
        df[date_col] = pd.to_datetime(df[date_col], errors="coerce")

# ── 10. Eliminar columnas 100% vacías ────────────────────────────────────────
before_cols = df.shape[1]
empty_cols = [c for c in df.columns if df[c].isna().all()]
print(f"  Columnas 100% vacías eliminadas: {empty_cols}")
df.drop(columns=empty_cols, inplace=True)
print(f"  Columnas finales: {df.shape[1]} (antes {before_cols})")

# ── 11. Eliminar duplicados exactos ──────────────────────────────────────────
dupes = df.duplicated().sum()
df.drop_duplicates(inplace=True)
df.reset_index(drop=True, inplace=True)
print(f"  Duplicados eliminados: {dupes}")
print(f"  Shape final: {df.shape}")

# ── 12. Reemplazar NaN con vacío para exportar limpio ────────────────────────
df_export = df.copy()
for col in df_export.select_dtypes(include="object").columns:
    df_export[col] = df_export[col].fillna("")

# ── 13. Exportar a nuevo Excel con estilo ─────────────────────────────────────
print("\nExportando Excel limpio...")
with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl", datetime_format="DD/MM/YYYY") as writer:
    df_export.to_excel(writer, sheet_name="BASE DE DATOS", index=False)
    ws = writer.sheets["BASE DE DATOS"]

    # Estilo de encabezado
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    ws.row_dimensions[1].height = 35

    # Ancho automático de columnas
    for col_idx, col in enumerate(df_export.columns, 1):
        max_len = max(
            len(str(col)),
            df_export[col].astype(str).str.len().max() if not df_export[col].empty else 0
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 30)

    # Filas alternadas
    alt_fill = PatternFill("solid", fgColor="EBF3FB")
    for row_idx in range(2, ws.max_row + 1):
        for cell in ws[row_idx]:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Congelar primera fila
    ws.freeze_panes = "A2"

    # Filtros automáticos
    ws.auto_filter.ref = ws.dimensions

# ── 14. Reporte de calidad ────────────────────────────────────────────────────
print("\n=== REPORTE DE CALIDAD ===")
print(f"Total filas limpias : {len(df_export):,}")
print(f"Total columnas      : {len(df_export.columns)}")
print("\nNulos por columna (%):")
null_pct = (df.isna().sum() / len(df) * 100).sort_values(ascending=False)
for col, pct in null_pct[null_pct > 0].items():
    print(f"  {col:<45} {pct:>5.1f}%")

print(f"\nArchivo guardado en: {OUTPUT_FILE}")
